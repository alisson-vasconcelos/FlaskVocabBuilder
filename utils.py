from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Definição dos lotes e valores
LOTE_3_CIDADES = [
    'Guará', 'Arniqueiras', 'Águas Claras', 'Park Way', 'Núcleo Bandeirante',
    'Candangolândia', 'SCIA/Estrutural', 'Vicente Pires', 'Riacho Fundo I',
    'Sobradinho', 'Sobradinho II', 'Fercal', 'Planaltina', 'Arapoanga',
    'Paranoá', 'Itapoã'
]

LOTE_5_CIDADES = [
    'Lago Sul', 'Jardim Botânico', 'São Sebastião', 'Brazlândia', 'Ceilândia',
    'Taguatinga', 'Sol Nascente/Por do Sol', 'Gama', 'Santa Maria',
    'Recanto das Emas', 'Água Quente', 'Samambaia', 'Riacho Fundo II'
]

VALOR_LOTE_3 = 575.75  # R$ por tonelada
VALOR_LOTE_5 = 591.82  # R$ por tonelada

def calcular_lote_e_valor(local_carga, quantidade_kg):
    """
    Calcula o lote e valor da carga baseado no local de carga e quantidade
    
    Args:
        local_carga (str): Local de carga
        quantidade_kg (float): Quantidade em quilos
    
    Returns:
        tuple: (lote, valor_carga) ou (None, None) se cidade não encontrada
    """
    # Normalizar o nome da cidade para comparação
    cidade_normalizada = local_carga.strip().title()
    
    # Verificar se pertence ao Lote 3
    for cidade in LOTE_3_CIDADES:
        if cidade.lower() in cidade_normalizada.lower() or cidade_normalizada.lower() in cidade.lower():
            toneladas = quantidade_kg / 1000
            valor_carga = toneladas * VALOR_LOTE_3
            return 3, round(valor_carga, 2)
    
    # Verificar se pertence ao Lote 5
    for cidade in LOTE_5_CIDADES:
        if cidade.lower() in cidade_normalizada.lower() or cidade_normalizada.lower() in cidade.lower():
            toneladas = quantidade_kg / 1000
            valor_carga = toneladas * VALOR_LOTE_5
            return 5, round(valor_carga, 2)
    
    # Cidade não encontrada
    return None, None

def gerar_ticket_pdf(pesagem, caminho_arquivo):
    """
    Gera um ticket de pesagem em PDF
    
    Args:
        pesagem: Objeto Pesagem com os dados
        caminho_arquivo (str): Caminho do arquivo PDF a ser gerado
    """
    doc = SimpleDocTemplate(caminho_arquivo, pagesize=A4)
    styles = getSampleStyleSheet()
    
    # Estilo customizado para título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=30
    )
    
    # Estilo para informações
    info_style = ParagraphStyle(
        'CustomInfo',
        parent=styles['Normal'],
        fontSize=12,
        alignment=TA_LEFT,
        spaceAfter=10
    )
    
    # Construir conteúdo do documento
    story = []
    
    # Título
    story.append(Paragraph("TICKET DE PESAGEM", title_style))
    story.append(Spacer(1, 20))
    
    # Informações em formato de tabela
    data = [
        ['Data:', pesagem.data.strftime('%d/%m/%Y')],
        ['Local da Carga:', pesagem.local_carga],
        ['Local da Descarga:', pesagem.local_descarga],
        ['Placa do Veículo:', pesagem.placa_veiculo],
        ['Motorista:', pesagem.motorista],
        ['Tipo de Produto:', pesagem.tipo_produto],
        ['Quantidade (Kg):', f'{pesagem.quantidade_kg:,.0f} kg'],
        ['Lote:', f'Lote {pesagem.lote}'],
        ['Valor da Carga:', f'R$ {pesagem.valor_carga:,.2f}']
    ]
    
    table = Table(data, colWidths=[2*inch, 4*inch])
    table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    story.append(table)
    story.append(Spacer(1, 30))
    
    # Rodapé
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        textColor=colors.grey
    )
    
    story.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", footer_style))
    
    # Construir PDF
    doc.build(story)

def gerar_relatorio_excel(pesagens, caminho_arquivo):
    """
    Gera um relatório em Excel com todas as pesagens
    
    Args:
        pesagens: Lista de objetos Pesagem
        caminho_arquivo (str): Caminho do arquivo Excel a ser gerado
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Pesagens"
    
    # Cabeçalhos
    headers = ['Placa', 'Motorista', 'Cidade (Local de Carga)', 'Lote', 'Data da Descarga', 'Kg (Quilos)', 'Valor da Carga']
    
    # Aplicar formatação aos cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dados
    for row, pesagem in enumerate(pesagens, 2):
        ws.cell(row=row, column=1, value=pesagem.placa_veiculo)
        ws.cell(row=row, column=2, value=pesagem.motorista)
        ws.cell(row=row, column=3, value=pesagem.local_carga)
        ws.cell(row=row, column=4, value=f'Lote {pesagem.lote}')
        ws.cell(row=row, column=5, value=pesagem.data.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=6, value=pesagem.quantidade_kg)
        ws.cell(row=row, column=7, value=f'R$ {pesagem.valor_carga:.2f}')
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Salvar arquivo
    wb.save(caminho_arquivo)

def obter_lista_cidades():
    """
    Retorna lista combinada de todas as cidades para uso em formulários
    
    Returns:
        list: Lista de todas as cidades ordenadas alfabeticamente
    """
    todas_cidades = LOTE_3_CIDADES + LOTE_5_CIDADES
    return sorted(todas_cidades)
